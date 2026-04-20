#!/usr/bin/env python3
"""
03_arch_schedule.py
Import the Architectural shop canonical schedule from Production Schedule_new.xlsx
and 1010 A shop Production Schedule.xlsx.

For each row:
  - Create/update a Work Order named JWM-ARCH-{job_id}
  - Create/update a JWM Production Schedule Line keyed by
    schedule_key = ARCH|{job_id}|{source_sheet}|{row_num}
"""
from __future__ import annotations
import os, json, pathlib, datetime
import openpyxl
from _frappe import FrappeClient, setup_logging, map_status_emoji

LOG = setup_logging("./logs", "03_arch_schedule")
ATT = pathlib.Path(os.environ["JWM_ATTACHMENTS_DIR"])

SOURCES = [
    ("Production Schedule_new.xlsx", "Production Schedule", "canonical"),
    ("1010 A shop Production Schedule.xlsx", "Production Schedule", "ashop"),
]


def to_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    try:
        return str(v)[:10]
    except Exception:
        return None


def row_status(row, idx):
    # In A-shop file some rows carry emoji status in a sub-sheet; in canonical
    # sheet we derive from Ship Late / Days On Hold. For demo we leave Unknown
    # unless there's a "Latest Comment" with a known keyword.
    return ("Unknown", "")


def main():
    fc = FrappeClient.from_env()

    wo_count = 0
    line_count = 0

    for fname, sheet, shop_tag in SOURCES:
        fpath = ATT / fname
        if not fpath.exists():
            LOG.warning("Missing: %s", fpath)
            continue
        LOG.info("Opening %s [%s]", fpath, sheet)
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        if sheet not in wb.sheetnames:
            LOG.error("Sheet %s not found in %s", sheet, fname)
            wb.close()
            continue
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        headers = [h.strip() if isinstance(h, str) else h for h in next(it)]
        idx = {h: i for i, h in enumerate(headers) if h}

        if "ID" not in idx:
            LOG.error("No ID column in %s/%s", fname, sheet)
            wb.close()
            continue

        row_num = 1
        for row in it:
            row_num += 1
            if not row or all(v is None for v in row):
                continue
            job_id = row[idx["ID"]]
            if not job_id:
                continue
            job_id = str(job_id).strip()
            job_name = str(row[idx["Job Name"]]).strip() if "Job Name" in idx and row[idx["Job Name"]] else ""

            ship_target = to_date(row[idx["Ship Target"]]) if "Ship Target" in idx else None
            ship_actual = to_date(row[idx["Ship Actual"]]) if "Ship Actual" in idx else None
            requested = to_date(row[idx["Requested Ship Date"]]) if "Requested Ship Date" in idx else None
            adjusted = to_date(row[idx["Adjusted Ship Date"]]) if "Adjusted Ship Date" in idx else None
            station = row[idx["Station"]] if "Station" in idx else None
            panel_qty = row[idx["Panel Qty"]] if "Panel Qty" in idx else None

            status_enum, status_raw = row_status(row, idx)

            raw = {
                headers[i]: (v.isoformat() if isinstance(v, (datetime.datetime, datetime.date)) else v)
                for i, v in enumerate(row)
                if i < len(headers) and headers[i]
            }

            # --- Work Order
            wo_name = f"JWM-ARCH-{job_id}"
            wo_payload = {
                "name": wo_name,
                "production_item": job_id,  # must match item_code from 01
                "qty": max(float(panel_qty), 1) if isinstance(panel_qty, (int, float)) and panel_qty else 1,
                "company": "JWM",
                "jwm_job_id": job_id,
                "jwm_shop": "Architectural",
                "jwm_status": status_enum,
                "jwm_raw_status": status_raw,
                "jwm_raw_data": json.dumps(raw, default=str)[:50000],
                "planned_start_date": to_date(row[idx.get("Release To Shop Target")]) if "Release To Shop Target" in idx else None,
                "expected_delivery_date": ship_target or adjusted or requested,
                "docstatus": 0,  # draft; we don't submit in import
            }
            # Phase 1: skip Work Order creation (requires BOM + fg_warehouse
            # per-item; Phase 2 will auto-generate BOMs). Schedule Line holds
            # the job data.
            wo_created = False

            # --- Schedule Line
            sk = f"ARCH|{job_id}|{sheet}|{row_num}"
            line_payload = {
                "schedule_key": sk,
                "job_id": job_id,
                "shop": "Architectural",
                "job_name": job_name,
                "status": status_enum,
                "jwm_raw_status": status_raw,
                "station": str(station)[:140] if station else None,
                "ship_target": ship_target,
                "ship_actual": ship_actual,
                "qty_required": float(panel_qty) if isinstance(panel_qty, (int, float)) else None,
                "source_file": fname,
                "source_sheet": sheet,
                "source_row": row_num,
                "work_order": wo_name if wo_created else None,
                "jwm_raw_data": json.dumps(raw, default=str)[:50000],
            }
            try:
                fc.upsert("JWM Production Schedule Line", line_payload, key_field="schedule_key")
                line_count += 1
            except Exception as e:
                LOG.error("Schedule Line %s failed: %s", sk, e)

        wb.close()

    LOG.info("Work Orders: %d  Schedule Lines: %d  Stats: %s", wo_count, line_count, fc.stats)
    print(json.dumps(fc.stats))


if __name__ == "__main__":
    main()
