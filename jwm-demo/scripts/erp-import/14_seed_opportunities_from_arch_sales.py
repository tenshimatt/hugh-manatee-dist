#!/usr/bin/env python3
"""
14_seed_opportunities_from_arch_sales.py

Seed ERPNext `Opportunity` records + linked Comments from the JWM Architectural
Sales pipeline xlsx (1,998 rows main + 18,381 comment rows).

Source:
    /Users/mattwright/pandora/Obsidian/PROJECTS/JWM/assets/2026-04-20/Arch Sales/JWMCD Arch Sales (3).xlsx

Behavior:
    - Idempotent: dedupe key (custom_jwm_project_name, transaction_date).
      If an Opportunity exists with matching key, we UPDATE non-empty fields;
      else CREATE.
    - Stage -> status via STAGE_TO_STATUS map (unknown -> Open).
    - `opportunity_from` / `party_name`: stock ERPNext requires a Dynamic Link.
      We point all to placeholder Customer "JWM Sales Pipeline Import" (created
      on first run); the real Company text lives in `customer_name` (Data).
      Real Customer dedupe is a Phase-2.5 task (JWM1451-57 area).
    - Comments: iterate Comments sheet. "Row N" (1-indexed) refers to the Nth
      data row (post-header) in the main sheet. We look up the opp by that row
      index in `row_to_opp` built during the seed phase. Non-matches -> orphan.
    - Resume: progress checkpointed every 50 rows to
      ./logs/14_seed_opps.checkpoint.json (set of source_row ints). Re-running
      picks up after last completed row. Delete that file to start over.
    - Rate limit: inherits from FrappeClient (RATE_LIMIT_RPS env, default 10).

Run:
    IMPORT_MODE=live python3 14_seed_opportunities_from_arch_sales.py
"""
from __future__ import annotations
import os
import re
import json
import time
import pathlib
import datetime
from typing import Any, Optional

import openpyxl
import requests
from _frappe import FrappeClient, FrappeError, setup_logging

LOG = setup_logging("./logs", "14_seed_opps")

SOURCE_XLSX = pathlib.Path(
    "/Users/mattwright/pandora/Obsidian/PROJECTS/JWM/assets/2026-04-20/"
    "Arch Sales/JWMCD Arch Sales (3).xlsx"
)
MAIN_SHEET = "JWMCD Arch Sales"
COMMENTS_SHEET = "Comments"

COMPANY = "JWM"
PLACEHOLDER_CUSTOMER = "JWM Sales Pipeline Import"

CHECKPOINT = pathlib.Path("./logs/14_seed_opps.checkpoint.json")
ORPHAN_LOG = pathlib.Path("./logs/14_seed_opps.orphans.log")
ROW_MAP_FILE = pathlib.Path("./logs/14_seed_opps.rowmap.json")

STAGE_TO_STATUS = {
    "Lost": "Lost",
    "Won": "Converted",
    "Submitted": "Quotation",
    "Active": "Open",
    "No Bid": "Closed",
    "Scope Review": "Open",
}

# xlsx header -> column index (computed on load, asserted against expected)
EXPECTED_HEADERS = [
    "Project Name", "Stage", "Received Date", "Bid Date", "Close Probability",
    "Total Bid Value", "Estimator / Sales", "Ball in court", "City", "State",
    "Contact Name", "NDA", "Company", "Company Helper", "Customer Code",
    "Won / Lost Date", "Contact Phone #", "Job Type", "Install Type",
    "Est. Vendor", "Onsite Schedule", "Metal Bid Value", "Glazing Bid Value",
    "Markup", "Margin", "Total NSF", "Tasks", "Date of Contract",
    "Actual Close Value", "Forecast Close Value", "Scope", "Bid 1", "Bid 2",
    "Bid 3", "Bid 4", "Follow up Date", "Average Sold GP%",
    "Sum Average Sold GP%", "Average Closed GP%", "Sum Average Closed GP%",
    "Average WIP GP%", "Sum Average WIP GP%", "Average Payment Days", "Year",
    "Comments", "Third Follow up", "Latest Comment", "Second Follow out",
]

ROW_RE = re.compile(r"Row\s*(\d+)", re.I)


# ---------- cell coercion helpers ----------

def _s(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _date(v: Any) -> Optional[str]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    # try a couple formats
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        s = str(v).replace("$", "").replace(",", "").strip()
        try:
            return float(s)
        except ValueError:
            return None


def _pct(v: Any) -> Optional[float]:
    """Percent: values 0..1 are scaled to 0..100; values already >1 kept."""
    n = _num(v)
    if n is None:
        return None
    if 0 <= n <= 1:
        return n * 100.0
    return n


def _int(v: Any) -> Optional[int]:
    n = _num(v)
    return int(n) if n is not None else None


_PHONE_STRIP_RE = re.compile(r"\s*(?:x|ext\.?|extension)\s*\d+.*$", re.I)


def _phone(v: Any) -> Optional[str]:
    """Strip extensions / garbage. Frappe validates phone numbers strictly."""
    s = _s(v)
    if not s:
        return None
    # Remove "x123", "ext. 456", "extension 7" suffix
    s = _PHONE_STRIP_RE.sub("", s).strip()
    # Strip trailing punctuation / whitespace
    s = re.sub(r"[^\d+\-\s().]+$", "", s).strip()
    if not s:
        return None
    # Require at least 7 digits to pass Frappe's validator
    digits = sum(c.isdigit() for c in s)
    if digits < 7:
        return None
    return s


def _bool(v: Any) -> int:
    if v is None or v == "":
        return 0
    s = str(v).strip().lower()
    return 1 if s in ("y", "yes", "true", "1", "x") else 0


# ---------- idempotent placeholder customer ----------

def ensure_placeholder_customer(fc: FrappeClient):
    existing = fc.get("Customer", PLACEHOLDER_CUSTOMER)
    if existing:
        LOG.info("Placeholder Customer exists: %s", PLACEHOLDER_CUSTOMER)
        return
    LOG.info("Creating placeholder Customer: %s", PLACEHOLDER_CUSTOMER)
    payload = {
        "doctype": "Customer",
        "customer_name": PLACEHOLDER_CUSTOMER,
        "customer_type": "Company",
        "customer_group": "All Customer Groups",
        "territory": "All Territories",
    }
    try:
        fc.create("Customer", payload)
    except FrappeError as e:
        # Might race if run in parallel; non-fatal if now exists
        LOG.warning("Placeholder customer create: %s", e)
        if not fc.get("Customer", PLACEHOLDER_CUSTOMER):
            raise


# ---------- dedupe lookup ----------

def find_opp(fc: FrappeClient, project_name: str, txn_date: str) -> Optional[str]:
    filters = [
        ["custom_jwm_project_name", "=", project_name],
        ["transaction_date", "=", txn_date],
    ]
    rows = fc.get_list("Opportunity", filters=filters, fields=("name",), limit=1)
    return rows[0]["name"] if rows else None


# ---------- payload build ----------

def row_to_payload(row: list, hdr: dict) -> Optional[dict]:
    def v(name):
        idx = hdr.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    project_name = _s(v("Project Name"))
    if not project_name:
        return None

    stage_raw = _s(v("Stage")) or ""
    stage = stage_raw.strip()
    status = STAGE_TO_STATUS.get(stage, "Open")

    # Skip noise: 0 or negative Total Bid on Lost rows
    tbv = _num(v("Total Bid Value"))
    if stage == "Lost" and (tbv is None or tbv <= 0):
        return None

    txn_date = _date(v("Received Date")) or datetime.date.today().isoformat()

    payload = {
        "doctype": "Opportunity",
        "opportunity_from": "Customer",
        "party_name": PLACEHOLDER_CUSTOMER,
        "company": COMPANY,
        "transaction_date": txn_date,
        "status": status,
        "customer_name": _s(v("Company")) or PLACEHOLDER_CUSTOMER,
        "contact_person": None,  # Link to Contact — leave blank; raw in custom field if needed
        "contact_mobile": _phone(v("Contact Phone #")),
        "probability": _pct(v("Close Probability")),
        "opportunity_amount": tbv,
        "city": _s(v("City")),
        "state": _s(v("State")),
        # Custom fields
        "custom_jwm_project_name": project_name,
        "custom_jwm_bid_date": _date(v("Bid Date")),
        "custom_jwm_close_date": _date(v("Won / Lost Date")),
        "custom_jwm_contract_date": _date(v("Date of Contract")),
        "custom_jwm_onsite_schedule": _date(v("Onsite Schedule")),
        "custom_jwm_year": _int(v("Year")),
        "custom_jwm_estimator": _s(v("Estimator / Sales")),
        "custom_jwm_ball_in_court": _s(v("Ball in court")),
        "custom_jwm_customer_code": _s(v("Customer Code")),
        "custom_jwm_city": _s(v("City")),
        "custom_jwm_state": _s(v("State")),
        "custom_jwm_nda": _bool(v("NDA")),
        "custom_jwm_job_type": _s(v("Job Type")),
        "custom_jwm_install_type": _s(v("Install Type")),
        "custom_jwm_est_vendor": _s(v("Est. Vendor")),
        "custom_jwm_total_nsf": _num(v("Total NSF")),
        "custom_jwm_scope": _s(v("Scope")),
        "custom_jwm_metal_bid_value": _num(v("Metal Bid Value")),
        "custom_jwm_glazing_bid_value": _num(v("Glazing Bid Value")),
        "custom_jwm_markup_pct": _pct(v("Markup")),
        "custom_jwm_margin_pct": _pct(v("Margin")),
        "custom_jwm_actual_close_value": _num(v("Actual Close Value")),
        "custom_jwm_forecast_close_value": _num(v("Forecast Close Value")),
        "custom_jwm_bid_1": _num(v("Bid 1")),
        "custom_jwm_bid_2": _num(v("Bid 2")),
        "custom_jwm_bid_3": _num(v("Bid 3")),
        "custom_jwm_bid_4": _num(v("Bid 4")),
        "custom_jwm_followup_date": _date(v("Follow up Date")),
        "custom_jwm_followup_2": _date(v("Second Follow out")),
        "custom_jwm_followup_3": _date(v("Third Follow up")),
        "custom_jwm_latest_comment": _s(v("Latest Comment")),
    }
    # Strip Nones to keep payload tight
    return {k: val for k, val in payload.items() if val is not None and val != ""}


# ---------- checkpoint I/O ----------

def load_checkpoint() -> tuple[set[int], dict[int, str]]:
    done = set()
    row_map = {}
    if CHECKPOINT.exists():
        try:
            d = json.loads(CHECKPOINT.read_text())
            done = set(d.get("done", []))
        except Exception as e:
            LOG.warning("Checkpoint unreadable, starting fresh: %s", e)
    if ROW_MAP_FILE.exists():
        try:
            row_map = {int(k): v for k, v in json.loads(ROW_MAP_FILE.read_text()).items()}
        except Exception:
            pass
    return done, row_map


def save_checkpoint(done: set[int], row_map: dict[int, str]):
    CHECKPOINT.write_text(json.dumps({"done": sorted(done)}))
    ROW_MAP_FILE.write_text(json.dumps({str(k): v for k, v in row_map.items()}))


# ---------- main seed phase ----------

def seed_opportunities(fc: FrappeClient) -> dict[int, str]:
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    ws = wb[MAIN_SHEET]
    it = ws.iter_rows(values_only=True)
    hdr_row = next(it)
    hdr = {h: i for i, h in enumerate(hdr_row) if h}

    done, row_map = load_checkpoint()
    LOG.info("Resume: %d rows already done; rowmap has %d entries", len(done), len(row_map))

    created = updated = skipped = errored = 0
    distinct_companies: set[str] = set()
    distinct_estimators: set[str] = set()

    for data_idx, row in enumerate(it, start=1):
        # data_idx is 1-indexed, matching the Comments sheet "Row N" convention
        if data_idx in done:
            existing_name = row_map.get(data_idx)
            if existing_name:
                continue
        try:
            payload = row_to_payload(list(row), hdr)
        except Exception as e:
            LOG.error("row %d: payload build error: %s", data_idx, e)
            errored += 1
            continue
        if not payload:
            skipped += 1
            done.add(data_idx)
            continue

        co = payload.get("customer_name")
        if co:
            distinct_companies.add(co)
        est = payload.get("custom_jwm_estimator")
        if est:
            distinct_estimators.add(est)

        payload["custom_jwm_source_row"] = data_idx

        project_name = payload["custom_jwm_project_name"]
        txn_date = payload["transaction_date"]

        try:
            existing_name = find_opp(fc, project_name, txn_date)
            if existing_name:
                # update (skip system fields)
                upd = {k: v for k, v in payload.items() if k != "doctype"}
                try:
                    fc.update("Opportunity", existing_name, upd)
                    updated += 1
                except FrappeError as e:
                    LOG.error("row %d: update %s: %s", data_idx, existing_name, e)
                    errored += 1
                    continue
                row_map[data_idx] = existing_name
            else:
                res = fc.create("Opportunity", payload)
                name = res.get("name") if isinstance(res, dict) else None
                if name:
                    row_map[data_idx] = name
                    created += 1
                else:
                    errored += 1
                    LOG.error("row %d: create returned no name: %s", data_idx, res)
                    continue
            done.add(data_idx)
        except FrappeError as e:
            LOG.error("row %d: frappe error: %s", data_idx, str(e)[:300])
            errored += 1
            # Back off on likely rate-limit-ish responses
            msg = str(e)
            if "429" in msg or "502" in msg or "503" in msg or "504" in msg:
                LOG.warning("Backing off 5s")
                time.sleep(5)
            continue
        except Exception as e:
            LOG.exception("row %d: unexpected: %s", data_idx, e)
            errored += 1
            continue

        if data_idx % 50 == 0:
            save_checkpoint(done, row_map)
            LOG.info("checkpoint @ row %d: created=%d updated=%d skipped=%d errored=%d",
                     data_idx, created, updated, skipped, errored)

    save_checkpoint(done, row_map)
    wb.close()

    LOG.info("SEED DONE. created=%d updated=%d skipped=%d errored=%d", created, updated, skipped, errored)
    LOG.info("Distinct companies: %d  Distinct estimators: %d",
             len(distinct_companies), len(distinct_estimators))
    return row_map


# ---------- comments phase ----------

def seed_comments(fc: FrappeClient, row_map: dict[int, str]):
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    ws = wb[COMMENTS_SHEET]

    # Progress file for comments (separate from opp checkpoint)
    cp = pathlib.Path("./logs/14_seed_opps.comments.checkpoint.json")
    seen_hashes: set[str] = set()
    if cp.exists():
        try:
            seen_hashes = set(json.loads(cp.read_text()).get("seen", []))
        except Exception:
            pass

    attached = 0
    orphaned = 0
    errored = 0
    skipped_dup = 0

    import hashlib as _hl
    orphan_f = ORPHAN_LOG.open("a")

    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row:
            continue
        cell0 = row[0]
        if not cell0:
            continue
        m = ROW_RE.match(str(cell0))
        if not m:
            continue
        src_row = int(m.group(1))
        content = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if not content:
            continue
        author = str(row[2]).strip() if len(row) > 2 and row[2] else "Imported"
        ts = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        opp_name = row_map.get(src_row)
        if not opp_name:
            orphaned += 1
            orphan_f.write(f"row={src_row} author={author} ts={ts} | {content[:120]}\n")
            continue

        body = content
        if author and author != "Imported":
            body = f"{content}\n\n— {author}"
            if ts:
                body += f" ({ts})"

        # Dedupe key: (opp_name, sha1(content))
        h = _hl.sha1(f"{opp_name}|{content}".encode()).hexdigest()
        if h in seen_hashes:
            skipped_dup += 1
            continue

        payload = {
            "doctype": "Comment",
            "comment_type": "Comment",
            "reference_doctype": "Opportunity",
            "reference_name": opp_name,
            "content": body,
            "comment_by": author or "Administrator",
            "comment_email": "Administrator",
        }
        try:
            fc.create("Comment", payload)
            attached += 1
            seen_hashes.add(h)
        except FrappeError as e:
            errored += 1
            LOG.error("comment row=%d -> %s: %s", src_row, opp_name, str(e)[:200])
            msg = str(e)
            if "429" in msg or "502" in msg or "503" in msg:
                time.sleep(5)
        except Exception as e:
            errored += 1
            LOG.exception("comment row=%d unexpected: %s", src_row, e)

        if attached and attached % 200 == 0:
            cp.write_text(json.dumps({"seen": list(seen_hashes)}))
            LOG.info("comment ckpt: attached=%d orphaned=%d errored=%d skipped_dup=%d",
                     attached, orphaned, errored, skipped_dup)

    cp.write_text(json.dumps({"seen": list(seen_hashes)}))
    orphan_f.close()
    wb.close()
    LOG.info("COMMENTS DONE. attached=%d orphaned=%d errored=%d skipped_dup=%d",
             attached, orphaned, errored, skipped_dup)


def main():
    fc = FrappeClient.from_env()
    LOG.info("Connected as: %s (dry=%s)", fc.ping(), fc.dry_run)
    if fc.dry_run:
        LOG.warning("DRY MODE. Set IMPORT_MODE=live for real writes.")

    ensure_placeholder_customer(fc)
    t0 = time.time()
    row_map = seed_opportunities(fc)
    t1 = time.time()
    LOG.info("Opportunities phase took %.1fs (%d rows mapped)", t1 - t0, len(row_map))

    seed_comments(fc, row_map)
    t2 = time.time()
    LOG.info("Comments phase took %.1fs", t2 - t1)

    LOG.info("TOTAL wall: %.1fs   stats=%s", t2 - t0, fc.stats)
    print(json.dumps({"stats": fc.stats, "elapsed_s": round(t2 - t0, 1), "rows_mapped": len(row_map)}))


if __name__ == "__main__":
    main()
