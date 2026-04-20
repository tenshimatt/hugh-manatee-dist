#!/usr/bin/env python3
"""
04_production_schedule_proc.py
Import Processing shop schedule from 1040 T Shop Production Schedule.xlsx.
Schema is very different from Arch shop: we pull from the 'PRODUCTION REPORT'
and 'JMW JOB OPERATIONS' sheets.

For PRODUCTION REPORT rows: create Work Order named JWM-PROC-{job_num}
For JMW JOB OPERATIONS rows: attach as Schedule Lines to corresponding WO
"""
from __future__ import annotations
import os, json, pathlib, datetime
import openpyxl
from _frappe import FrappeClient, setup_logging

LOG = setup_logging("./logs", "04_production_schedule_proc")
ATT = pathlib.Path(os.environ["JWM_ATTACHMENTS_DIR"])
SRC = ATT / "1040 T Shop Production Schedule.xlsx"
ITEM_GROUP = "JWM Jobs"


def to_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v)[:10]


def main():
    fc = FrappeClient.from_env()
    if not SRC.exists():
        LOG.error("Missing %s", SRC)
        return
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

    # --- PRODUCTION REPORT -> WO + Item
    wo_count = 0
    if "PRODUCTION REPORT" in wb.sheetnames:
        ws = wb["PRODUCTION REPORT"]
        it = ws.iter_rows(values_only=True)
        headers = [str(h).replace("\n", " ").strip() if h else None for h in next(it)]
        idx = {h: i for i, h in enumerate(headers) if h}

        for rn, row in enumerate(it, start=2):
            if not row or all(v is None for v in row):
                continue
            job_num = row[idx.get("Job Num")] if "Job Num" in idx else None
            if not job_num:
                continue
            job_num = str(job_num).strip()
            cust_id = row[idx.get("Cust ID")] if "Cust ID" in idx else None
            part_num = row[idx.get("Part Num")] if "Part Num" in idx else None
            line_desc = row[idx.get("Line Desc")] if "Line Desc" in idx else ""
            order_qty = row[idx.get("Order Qty")] if "Order Qty" in idx else None
            ship_by = to_date(row[idx.get("Ship By")]) if "Ship By" in idx else None

            raw = {headers[i]: v for i, v in enumerate(row) if i < len(headers) and headers[i]}

            # Ensure Item
            item_code = f"PROC-{job_num}"
            try:
                fc.upsert(
                    "Item",
                    {
                        "item_code": item_code,
                        "item_name": f"{job_num} {part_num or ''}".strip()[:140],
                        "item_group": ITEM_GROUP,
                        "description": str(line_desc)[:2000],
                        "is_stock_item": 0,
                        "include_item_in_manufacturing": 1,
                        "stock_uom": "Nos",
                    },
                    key_field="item_code",
                )
            except Exception as e:
                LOG.error("Item %s failed: %s", item_code, e)

            # Ensure customer (Cust ID used as name; real customer lookup Phase 2)
            if cust_id:
                try:
                    fc.upsert(
                        "Customer",
                        {
                            "customer_name": str(cust_id),
                            "customer_type": "Company",
                            "customer_group": "Commercial",
                            "territory": "All Territories",
                        },
                        key_field="customer_name",
                    )
                except Exception as e:
                    LOG.warning("Customer %s: %s", cust_id, e)

            wo_name = f"JWM-PROC-{job_num}"
            wo_payload = {
                "name": wo_name,
                "production_item": item_code,
                "qty": max(float(order_qty), 1) if isinstance(order_qty, (int, float)) and order_qty else 1,
                "company": "JWM",
                "jwm_job_id": job_num,
                "jwm_shop": "Processing",
                "jwm_status": "Unknown",
                "jwm_raw_data": json.dumps(raw, default=str)[:50000],
                "expected_delivery_date": ship_by,
                "docstatus": 0,
            }
            # Phase 1: skip Work Order (requires BOM/warehouse). Phase 2 TODO.
            wo_created = False

            # Schedule line
            sk = f"PROC|{job_num}|PRODUCTION REPORT|{rn}"
            try:
                fc.upsert(
                    "JWM Production Schedule Line",
                    {
                        "schedule_key": sk,
                        "job_id": job_num,
                        "shop": "Processing",
                        "job_name": str(part_num or "")[:140],
                        "status": "Unknown",
                        "ship_target": ship_by,
                        "qty_required": float(order_qty) if isinstance(order_qty, (int, float)) else None,
                        "source_file": SRC.name,
                        "source_sheet": "PRODUCTION REPORT",
                        "source_row": rn,
                        "work_order": wo_name if wo_created else None,
                        "jwm_raw_data": json.dumps(raw, default=str)[:50000],
                    },
                    key_field="schedule_key",
                )
            except Exception as e:
                LOG.error("Sched line %s: %s", sk, e)

    # --- JMW JOB OPERATIONS -> Schedule Lines
    op_count = 0
    if "JMW JOB OPERATIONS" in wb.sheetnames:
        ws = wb["JMW JOB OPERATIONS"]
        it = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h else None for h in next(it)]
        idx = {h: i for i, h in enumerate(headers) if h}
        for rn, row in enumerate(it, start=2):
            if not row or all(v is None for v in row):
                continue
            jn = row[idx.get("Job Number")] if "Job Number" in idx else None
            if not jn:
                continue
            jn = str(jn).strip()
            op = row[idx.get("Operation")] if "Operation" in idx else None
            due = to_date(row[idx.get("Due Date")]) if "Due Date" in idx else None
            qty_req = row[idx.get("Qty Required")] if "Qty Required" in idx else None
            qty_cmp = row[idx.get("Qty Completed")] if "Qty Completed" in idx else None
            qty_rem = row[idx.get("Qty Remain")] if "Qty Remain" in idx else None
            hrs = row[idx.get("Est. Prod Hours")] if "Est. Prod Hours" in idx else None

            raw = {headers[i]: v for i, v in enumerate(row) if i < len(headers) and headers[i]}
            sk = f"PROC|{jn}|OPS|{rn}"
            wo_name = f"JWM-PROC-{jn.split('-')[0]}" if "-" in jn else f"JWM-PROC-{jn}"
            try:
                fc.upsert(
                    "JWM Production Schedule Line",
                    {
                        "schedule_key": sk,
                        "job_id": jn,
                        "shop": "Processing",
                        "status": "Unknown",
                        "station": str(op)[:140] if op else None,
                        "ship_target": due,
                        "qty_required": float(qty_req) if isinstance(qty_req, (int, float)) else None,
                        "qty_completed": float(qty_cmp) if isinstance(qty_cmp, (int, float)) else None,
                        "qty_remain": float(qty_rem) if isinstance(qty_rem, (int, float)) else None,
                        "est_hours": float(hrs) if isinstance(hrs, (int, float)) else None,
                        "source_file": SRC.name,
                        "source_sheet": "JMW JOB OPERATIONS",
                        "source_row": rn,
                        "jwm_raw_data": json.dumps(raw, default=str)[:50000],
                    },
                    key_field="schedule_key",
                )
                op_count += 1
            except Exception as e:
                LOG.error("Op line %s: %s", sk, e)

    wb.close()
    LOG.info("Proc Work Orders: %d  Op Lines: %d  Stats: %s", wo_count, op_count, fc.stats)
    print(json.dumps(fc.stats))


if __name__ == "__main__":
    main()
