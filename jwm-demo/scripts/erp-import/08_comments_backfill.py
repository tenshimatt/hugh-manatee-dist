#!/usr/bin/env python3
"""
08_comments_backfill.py
Pull the 'Comments' sheet from A-shop and T-shop workbooks; attach each comment
to the corresponding Work Order via a Comment docs.

Comment cell format (4 cols):
  ['Row N', '@user @user message', 'Author', 'MM/DD/YY HH:MM AM/PM']

Row N refers to the row number in the Production Schedule sheet of the same
workbook. We build row_num -> WO name map at import time.
"""
from __future__ import annotations
import os, re, json, pathlib, datetime
import openpyxl
from _frappe import FrappeClient, setup_logging

LOG = setup_logging("./logs", "08_comments_backfill")
ATT = pathlib.Path(os.environ["JWM_ATTACHMENTS_DIR"])

SOURCES = [
    ("1010 A shop Production Schedule.xlsx", "Production Schedule", "Comments", "Architectural"),
    # T-shop comments sheet doesn't follow the same Row-N convention; skip.
]

ROW_RE = re.compile(r"Row\s*(\d+)", re.I)


def build_row_map(ws, shop: str) -> dict[int, str]:
    """Returns row_num -> job_id (used to find Schedule Lines by job_id)."""
    it = ws.iter_rows(values_only=True)
    headers = [h.strip() if isinstance(h, str) else h for h in next(it)]
    idx = {h: i for i, h in enumerate(headers) if h}
    if "ID" not in idx:
        return {}
    out = {}
    rn = 1
    for row in it:
        rn += 1
        if not row:
            continue
        jid = row[idx["ID"]]
        if not jid:
            continue
        out[rn] = str(jid).strip()
    return out


def main():
    fc = FrappeClient.from_env()
    total = 0
    orphans = 0

    for fname, sched_sheet, comments_sheet, shop in SOURCES:
        fpath = ATT / fname
        if not fpath.exists():
            LOG.warning("Missing %s", fpath); continue
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        if sched_sheet not in wb.sheetnames or comments_sheet not in wb.sheetnames:
            wb.close(); continue
        row_map = build_row_map(wb[sched_sheet], shop)
        max_row = max(row_map.keys()) if row_map else 0
        LOG.info("%s: %d rows mapped (max row %d)", fname, len(row_map), max_row)

        ws = wb[comments_sheet]
        for rn, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not row:
                continue
            cell0 = row[0]
            if not cell0:
                continue
            m = ROW_RE.match(str(cell0))
            if not m:
                continue
            src_row = int(m.group(1))
            content = str(row[1]) if len(row) > 1 and row[1] else ""
            author = str(row[2]) if len(row) > 2 and row[2] else "Imported"
            ts = str(row[3]) if len(row) > 3 and row[3] else ""

            job_id = row_map.get(src_row)
            if not job_id:
                orphans += 1
                LOG.warning("Orphan comment Row %d (max %d): %s", src_row, max_row, content[:60])
                continue

            # Look up the Schedule Line for this job
            sched = fc.get_list(
                "JWM Production Schedule Line",
                filters=[["job_id", "=", job_id], ["shop", "=", shop]],
                fields=("name",),
                limit=1,
            )
            if not sched:
                orphans += 1
                LOG.warning("No Schedule Line for job %s", job_id)
                continue
            target_name = sched[0]["name"]

            body = f"{content}\n\n— {author} ({ts})"
            payload = {
                "comment_type": "Comment",
                "reference_doctype": "JWM Production Schedule Line",
                "reference_name": target_name,
                "content": body,
                "comment_by": "Administrator",
                "comment_email": "Administrator",
            }
            try:
                existing = fc.get_list(
                    "Comment",
                    filters=[
                        ["reference_doctype", "=", "JWM Production Schedule Line"],
                        ["reference_name", "=", target_name],
                        ["content", "like", body[:40] + "%"],
                    ],
                    fields=("name",),
                    limit=1,
                )
                if existing:
                    continue
                if fc.dry_run:
                    LOG.info("[DRY] Comment -> %s: %s", target_name, body[:80])
                    fc.stats["dry_create"] += 1
                else:
                    fc.create("Comment", payload)
                total += 1
            except Exception as e:
                LOG.error("Comment row %d -> %s: %s", src_row, target_name, e)
        wb.close()

    LOG.info("Comments: %d   Orphans: %d   Stats: %s", total, orphans, fc.stats)
    print(json.dumps(fc.stats))


if __name__ == "__main__":
    main()
