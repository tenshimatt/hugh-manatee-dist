#!/usr/bin/env python3
"""
05_daily_efficiency.py
Import Daily Efficiency Log.xlsx into JWM Daily Efficiency DocType.
Each station (sheet) has per-day rows; header is in row 2 (row 1 is a section header).
"""
from __future__ import annotations
import os, json, pathlib, datetime
import openpyxl
from _frappe import FrappeClient, setup_logging

LOG = setup_logging("./logs", "05_daily_efficiency")
ATT = pathlib.Path(os.environ["JWM_ATTACHMENTS_DIR"])
SRC = ATT / "Daily Efficiency Log.xlsx"

SKIP_SHEETS = {"Data", "Downtime"}


def norm_time(v):
    if v is None:
        return None
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, datetime.datetime):
        return v.strftime("%H:%M:%S")
    return None


def norm_date(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    return str(v)[:10]


def main():
    fc = FrappeClient.from_env()
    if not SRC.exists():
        LOG.error("Missing %s", SRC)
        return
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

    total = 0
    for sheet in wb.sheetnames:
        if sheet in SKIP_SHEETS:
            continue
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        try:
            _ = next(it)          # section header row (Estimate/Actual/Setup/Results)
            header_row = next(it) # real headers
        except StopIteration:
            continue
        headers = [str(h).strip() if h else None for h in header_row]
        idx = {h: i for i, h in enumerate(headers) if h}
        if "Date" not in idx or "Job#" not in idx:
            LOG.warning("Sheet %s: no Date/Job# columns, skip", sheet)
            continue

        for rn, row in enumerate(it, start=3):
            if not row or all(v is None for v in row):
                continue
            dt = norm_date(row[idx["Date"]])
            if not dt:
                continue
            job_num = row[idx["Job#"]]
            if not job_num:
                continue
            key = f"{dt}-{sheet}-1st"  # shift 1 default; Daily Efficiency log doesn't split
            payload = {
                "efficiency_key": key,
                "log_date": dt,
                "workstation": sheet,
                "shift": "1st",
                "job_number": str(job_num)[:140],
                "part_number": str(row[idx["Part#"]])[:140] if "Part#" in idx and row[idx["Part#"]] else None,
                "job_type": str(row[idx["Job Type"]])[:80] if "Job Type" in idx and row[idx["Job Type"]] else None,
                "start_time": norm_time(row[idx["Start Time"]]) if "Start Time" in idx else None,
                "end_time": norm_time(row[idx["End Time"]]) if "End Time" in idx else None,
                "qty_required": float(row[idx["Qty Required"]]) if "Qty Required" in idx and isinstance(row[idx["Qty Required"]], (int, float)) else None,
                "qty_complete": float(row[idx["Qty Complete"]]) if "Qty Complete" in idx and isinstance(row[idx["Qty Complete"]], (int, float)) else None,
                "source_file": SRC.name,
                "source_sheet": sheet,
                "jwm_raw_data": json.dumps(
                    {headers[i]: v for i, v in enumerate(row) if i < len(headers) and headers[i]},
                    default=str,
                )[:50000],
            }
            try:
                # Unique constraint on efficiency_key — we must dedup even if
                # key accidentally repeats across jobs same day/workstation.
                # Append short job hash for uniqueness.
                import hashlib
                payload["efficiency_key"] = f"{key}-{hashlib.md5(str(job_num).encode()).hexdigest()[:6]}-{rn}"
                fc.upsert(
                    "JWM Daily Efficiency", payload, key_field="efficiency_key"
                )
                total += 1
            except Exception as e:
                LOG.error("DE %s row %d: %s", sheet, rn, e)

    wb.close()
    LOG.info("Daily Efficiency rows: %d  Stats: %s", total, fc.stats)
    print(json.dumps(fc.stats))


if __name__ == "__main__":
    main()
