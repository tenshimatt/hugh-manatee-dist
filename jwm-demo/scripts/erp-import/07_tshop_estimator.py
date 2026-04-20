#!/usr/bin/env python3
"""
07_tshop_estimator.py
Tshop Estimator Update R18.xlsx — reference data for estimating.
SCOPE (phase 1): Import the 'FB Data' and 'TL Data' sheets as Items (material
reference), and ingest 'Phase Codes' as a reference table via Custom DocType
or as notes on Items.

Phase 2 TODO: parse Estimator logic itself (formulas) into BOMs / Operations.
"""
from __future__ import annotations
import os, json, pathlib
import openpyxl
from _frappe import FrappeClient, setup_logging

LOG = setup_logging("./logs", "07_tshop_estimator")
ATT = pathlib.Path(os.environ["JWM_ATTACHMENTS_DIR"])
SRC = ATT / "Tshop Estimator Update R18.xlsx"
ITEM_GROUP = "JWM Raw Materials"


def ensure_group(fc):
    fc.upsert(
        "Item Group",
        {
            "item_group_name": ITEM_GROUP,
            "parent_item_group": "All Item Groups",
            "is_group": 0,
        },
        key_field="item_group_name",
    )


def main():
    fc = FrappeClient.from_env()
    if not SRC.exists():
        LOG.error("Missing %s", SRC)
        return
    ensure_group(fc)
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

    total = 0
    for sheet in ("FB Data", "TL Data"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h else None for h in next(it)]
        idx = {h: i for i, h in enumerate(headers) if h}
        if "Item ID" not in idx or "Name" not in idx:
            continue
        for rn, row in enumerate(it, start=2):
            if not row or all(v is None for v in row):
                continue
            iid = row[idx["Item ID"]]
            if not iid:
                continue
            name = row[idx["Name"]]
            uw = row[idx.get("Unit Weight")] if "Unit Weight" in idx else None
            raw = {headers[i]: v for i, v in enumerate(row) if i < len(headers) and headers[i]}
            try:
                fc.upsert(
                    "Item",
                    {
                        "item_code": str(iid)[:140],
                        "item_name": str(name)[:140] if name else str(iid)[:140],
                        "item_group": ITEM_GROUP,
                        "stock_uom": "Nos",
                        "is_stock_item": 1,
                        "description": f"Unit Weight: {uw}\n{json.dumps(raw, default=str)}"[:5000],
                    },
                    key_field="item_code",
                )
                total += 1
            except Exception as e:
                LOG.error("Item %s: %s", iid, e)

    wb.close()
    LOG.info("Estimator materials: %d  Stats: %s", total, fc.stats)
    print(json.dumps(fc.stats))


if __name__ == "__main__":
    main()
