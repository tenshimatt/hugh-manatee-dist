#!/usr/bin/env python3
"""
09_status_derivation.py

Phase-2 enrichment: derive and patch `jwm_stage`, `jwm_status_emoji`,
`jwm_current_station`, and `jwm_priority` onto every existing
`JWM Production Schedule Line`.

Source of truth:  /Users/mattwright/pandora/Obsidian/PROJECTS/JWM/50-research/attachments/1010 A shop Production Schedule.xlsx
  - "Production Schedule" sheet (main) supplies "By" columns for stage
    derivation + `Ranked Priority` + `Station` + `Ship Target`.
  - Per-station sheets (AXYZ Titans CNC, AXYZ CNC, Clean and Brake, Fab,
    Shear, Cidan, Extrusion Saw, Punch, Roller, Ship) give per-job emoji
    status + current station name.

Idempotent: compares existing document values to the derived payload and
only PATCHes when something actually changed.

Run:
    cd /Users/mattwright/pandora/jwm-demo/scripts/erp-import
    source .env       # FRAPPE_* + IMPORT_MODE=live for the real run
    python3 09_status_derivation.py
"""
from __future__ import annotations

import datetime as _dt
import json
import logging
import os
import pathlib
import sys
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl

from _frappe import FrappeClient, FrappeError, setup_logging, map_status_emoji

SCRIPT_NAME = "09_status_derivation"
LOG_DIR = pathlib.Path(os.getenv("LOG_DIR", "./logs")).resolve()
TMP_LOG_DIR = pathlib.Path("/tmp/jwm-import")
TMP_LOG_DIR.mkdir(parents=True, exist_ok=True)

LOG = setup_logging(str(LOG_DIR), SCRIPT_NAME)
# Mirror to /tmp path for quick tailing.
_tmp_handler = logging.FileHandler(TMP_LOG_DIR / f"{SCRIPT_NAME}.log", mode="a")
_tmp_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
LOG.addHandler(_tmp_handler)

# ------------------------------ source file ---------------------------------
ATTACH_CANDIDATES = [
    "/Users/mattwright/pandora/Obsidian/PROJECTS/JWM/50-research/attachments/1010 A shop Production Schedule.xlsx",
    "/Users/mattwright/pandora/Obsidian/PROJECTS/JWM/50-research/attachments/1010 A shop Production Schedule.xlsx",
]

# ------------------------------ canonical stages ----------------------------
# Canonical stage order (earliest -> latest). Used to pick the "furthest right
# non-empty" By column so jobs surface at their *current* stage.
STAGE_ORDER = [
    "uncategorized",
    "evaluating",
    "float",
    "layout",
    "layout_check",
    "sketch",
    "sketch_check",
    "correction",
    "cnc_prog",
    "laser_prog",
    "punch_prog",
    "prog_complete",
    "release_shop",
]

# Spreadsheet "By" columns mapped to canonical stage slugs.
BY_COLUMN_TO_STAGE = [
    # (spreadsheet header, canonical stage slug)
    ("Evaluation By", "evaluating"),
    ("LO By", "layout"),
    ("LO Check By", "layout_check"),
    ("Sketch By", "sketch"),
    ("Sketch Check By", "sketch_check"),
    ("Correction By", "correction"),
    ("AXYZ Prog By", "cnc_prog"),
    ("Laser Prog By", "laser_prog"),
    ("Punch Prog By", "punch_prog"),
    # `Prog Complete` is a date column, not a "By" column, but it gates
    # the prog_complete stage.  Present in the xlsx header.
    ("Prog Complete", "prog_complete"),
    ("Release By", "release_shop"),
]
STAGE_INDEX = {s: i for i, s in enumerate(STAGE_ORDER)}

# ------------------------------ station sheets ------------------------------
# Ordered from MOST DOWNSTREAM (ship) to MOST UPSTREAM.  If a job appears in
# multiple sub-sheets we believe the furthest-downstream one.
STATION_SHEETS: List[Tuple[str, str]] = [
    # (sheet name, default station label)
    # Real shop-floor stations, most-downstream first --- the first sheet that
    # claims a job wins.  "Ship Schedule" is a summary (not a station) and is
    # evaluated LAST as a fallback for jobs not claimed by any actual station.
    ("Roller Schedule", "Roller"),
    ("Punch Schedule", "Punch"),
    ("Extrusion Saw Schedule", "Extrusion Saw"),
    ("Cidan Schedule", "Cidan"),
    ("Shear Schedule", "Shear"),
    ("Fab Schedule", "Fab"),
    ("Clean and Brake Schedule", "Clean and Brake"),
    ("AXYZ CNC Schedule", "AXYZ CNC"),
    ("AXYZ Titans CNC Schedule", "AXYZ Titans CNC"),
    ("Ship Schedule", "Ship"),
]


# ---------------------------- helpers ---------------------------------------
def _norm_job_id(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.lower() == "none":
        return None
    # collapse whitespace and drop trailing ".0" from numeric ids
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _any_truthy(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, str):
        return v.strip() != ""
    return True


def _extract_emoji(*cells: Any) -> Optional[str]:
    """Scan cells for a known emoji or status keyword; return canonical emoji."""
    for c in cells:
        if c is None:
            continue
        s = str(c)
        if "🟢" in s or "ON TIME" in s.upper() or "ON TRACK" in s.upper():
            return "🟢"
        if "🔴" in s or "LATE" in s.upper() or "OVERDUE" in s.upper():
            return "🔴"
        if "🟡" in s or "WARNING" in s.upper():
            return "🟡"
        if "❌" in s or "NOT FOUND" in s.upper():
            return "❌"
    return None


def _derive_stage(by_values: Dict[str, Any]) -> str:
    """Walk canonical By columns from LAST -> FIRST; return first populated."""
    # Walk in reverse (latest stage first); first populated wins.
    for header, slug in reversed(BY_COLUMN_TO_STAGE):
        if _any_truthy(by_values.get(header)):
            return slug
    return "uncategorized"


def _derive_priority_from_rank(rank: Any) -> Optional[str]:
    try:
        r = float(rank)
    except (TypeError, ValueError):
        return None
    if r <= 10:
        return "High"
    if r <= 30:
        return "Medium"
    return "Low"


def _derive_priority_from_ship(ship_target: Any, today: _dt.date) -> Optional[str]:
    if ship_target is None:
        return None
    if isinstance(ship_target, _dt.datetime):
        d = ship_target.date()
    elif isinstance(ship_target, _dt.date):
        d = ship_target
    else:
        try:
            d = _dt.datetime.fromisoformat(str(ship_target)[:10]).date()
        except Exception:
            return None
    delta = (d - today).days
    if delta <= 7:
        return "High"
    if delta <= 21:
        return "Medium"
    return "Low"


# ------------------------- build enrichment map -----------------------------
def build_enrichment(xlsx_path: pathlib.Path) -> Dict[str, Dict[str, Any]]:
    """Return job_id -> {jwm_stage, jwm_status_emoji, jwm_current_station,
    jwm_priority, _sources: [...]}."""
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)

    enrich: Dict[str, Dict[str, Any]] = {}
    multi_claim: Dict[str, List[str]] = {}
    today = _dt.date.today()

    # --- 1. MAIN SHEET: stage + priority + fallback station ------------------
    if "Production Schedule" not in wb.sheetnames:
        raise RuntimeError("Main 'Production Schedule' sheet missing from xlsx")
    ws = wb["Production Schedule"]
    it = ws.iter_rows(values_only=True)
    headers = list(next(it))
    hdr_idx = {h: i for i, h in enumerate(headers) if h}

    id_col = hdr_idx.get("ID", 1)
    rank_col = hdr_idx.get("Ranked Priority")
    ship_col = hdr_idx.get("Ship Target")
    station_col = hdr_idx.get("Station")

    by_col_idx = [
        (h, slug, hdr_idx[h]) for (h, slug) in BY_COLUMN_TO_STAGE if h in hdr_idx
    ]

    for row in it:
        if not row or row[id_col] is None:
            continue
        jid = _norm_job_id(row[id_col])
        if not jid:
            continue
        by_values = {h: row[idx] for (h, _slug, idx) in by_col_idx}
        stage = _derive_stage(by_values)

        prio = None
        if rank_col is not None:
            prio = _derive_priority_from_rank(row[rank_col])
        if prio is None and ship_col is not None:
            prio = _derive_priority_from_ship(row[ship_col], today)
        if prio is None:
            prio = "Low"

        station_fallback = None
        if station_col is not None and _any_truthy(row[station_col]):
            station_fallback = str(row[station_col]).strip()

        enrich[jid] = {
            "jwm_stage": stage,
            "jwm_status_emoji": "⚪",
            "jwm_current_station": station_fallback or "",
            "jwm_priority": prio,
            "_sources": ["Production Schedule"],
        }

    LOG.info("Main sheet: enrichment rows = %d", len(enrich))

    # --- 2. STATION SUB-SHEETS: emoji + station (most-downstream wins) -------
    for sheet_name, default_label in STATION_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 7:
            continue

        # Row 5 (0-indexed) is the data-header row in station sheets.
        data_header = rows[5] if len(rows) > 5 else ()
        sub_hdr_idx = {
            str(h).strip(): i for i, h in enumerate(data_header) if h
        }
        id_c = sub_hdr_idx.get("ID")
        job_id_c = sub_hdr_idx.get("Job ID")  # Ship Schedule
        cur_st_c = sub_hdr_idx.get("Current Station")
        queue_c = sub_hdr_idx.get("Queue Position")

        # Ship Schedule row layout differs: header row is row 3, data from row 4.
        is_ship_sheet = sheet_name == "Ship Schedule"
        if is_ship_sheet:
            data_header = rows[3]
            sub_hdr_idx = {
                str(h).strip(): i for i, h in enumerate(data_header) if h
            }
            id_c = sub_hdr_idx.get("Job ID")
            cur_st_c = sub_hdr_idx.get("Current Station")
            queue_c = None
            data_start = 4
        else:
            data_start = 6

        id_col_use = id_c if id_c is not None else job_id_c
        if id_col_use is None:
            LOG.warning("Sub-sheet %s: no ID column found; skipping", sheet_name)
            continue

        for r_i, row in enumerate(rows[data_start:], start=data_start):
            if not row:
                continue
            if id_col_use >= len(row):
                continue
            jid = _norm_job_id(row[id_col_use])
            if not jid:
                continue
            # Skip header-echo rows like "ID"
            if jid.upper() == "ID":
                continue

            # Emoji may live in the row itself (some station sheets encode a
            # status col), but primarily we rely on the "SEARCH JOB ID" box
            # at row 0 --- which only carries emoji for ONE job.  So for most
            # rows we infer emoji from Ship Target vs today.
            emoji = _extract_emoji(*[c for c in row])

            # Inferred emoji from ship target when not explicit
            if not emoji:
                ship_idx = sub_hdr_idx.get("Ship Target")
                if ship_idx is not None and ship_idx < len(row):
                    sv = row[ship_idx]
                    if isinstance(sv, (_dt.datetime, _dt.date)):
                        d = sv.date() if isinstance(sv, _dt.datetime) else sv
                        days = (d - today).days
                        if days < 0:
                            emoji = "🔴"
                        elif days <= 7:
                            emoji = "🟡"
                        else:
                            emoji = "🟢"

            if not emoji:
                emoji = "⚪"

            # Station label
            station_label = default_label
            if cur_st_c is not None and cur_st_c < len(row) and _any_truthy(row[cur_st_c]):
                station_label = str(row[cur_st_c]).strip()
            elif queue_c is not None and queue_c < len(row) and _any_truthy(row[queue_c]):
                qv = str(row[queue_c]).strip()
                # Queue position in AXYZ CNC sheet is actually a stage indicator
                # (LO, Float, Sketch...) --- use it only if it looks station-ish.
                station_label = f"{default_label} ({qv})"

            if jid not in enrich:
                # Job not in main sheet --- create a bare record.
                enrich[jid] = {
                    "jwm_stage": "uncategorized",
                    "jwm_status_emoji": emoji,
                    "jwm_current_station": station_label,
                    "jwm_priority": "Low",
                    "_sources": [sheet_name],
                }
                continue

            existing_src = enrich[jid]["_sources"]
            if sheet_name in existing_src:
                continue
            # Most-downstream sheet wins; STATION_SHEETS iterates in that order
            # so the FIRST claim we see is already the most-downstream one.
            if len(existing_src) == 1 and existing_src[0] == "Production Schedule":
                enrich[jid]["jwm_status_emoji"] = emoji
                enrich[jid]["jwm_current_station"] = station_label
                enrich[jid]["_sources"].append(sheet_name)
            else:
                # Already had a downstream claim; record it for reporting but
                # don't overwrite.
                multi_claim.setdefault(jid, [existing_src[-1]]).append(sheet_name)

    wb.close()

    LOG.info("Multi-station claims: %d jobs", len(multi_claim))
    for jid, srcs in list(multi_claim.items())[:10]:
        LOG.info("  %s claimed by %s (kept %s)", jid, srcs, enrich[jid]["jwm_current_station"])

    return enrich


# ---------------------- ensure custom fields exist --------------------------
DT_TARGET = "JWM Production Schedule Line"
MODULE = "Custom"

STAGE_OPTIONS = "\n".join(STAGE_ORDER)

CUSTOM_FIELDS = [
    {
        "doctype": "Custom Field",
        "dt": DT_TARGET,
        "fieldname": "jwm_stage",
        "label": "JWM Stage",
        "fieldtype": "Select",
        "options": STAGE_OPTIONS,
        "default": "uncategorized",
        "insert_after": "station",
        "in_list_view": 1,
        "in_standard_filter": 1,
        "module": MODULE,
    },
    {
        "doctype": "Custom Field",
        "dt": DT_TARGET,
        "fieldname": "jwm_status_emoji",
        "label": "JWM Status Emoji",
        "fieldtype": "Data",
        "insert_after": "jwm_stage",
        "in_list_view": 1,
        "module": MODULE,
    },
    {
        "doctype": "Custom Field",
        "dt": DT_TARGET,
        "fieldname": "jwm_current_station",
        "label": "JWM Current Station",
        "fieldtype": "Data",
        "insert_after": "jwm_status_emoji",
        "in_list_view": 1,
        "module": MODULE,
    },
    {
        "doctype": "Custom Field",
        "dt": DT_TARGET,
        "fieldname": "jwm_priority",
        "label": "JWM Priority",
        "fieldtype": "Select",
        "options": "High\nMedium\nLow",
        "default": "Low",
        "insert_after": "jwm_current_station",
        "in_list_view": 1,
        "in_standard_filter": 1,
        "module": MODULE,
    },
]


def ensure_custom_fields(fc: FrappeClient) -> None:
    for cf in CUSTOM_FIELDS:
        fq = f'{cf["dt"]}-{cf["fieldname"]}'
        existing = fc.get("Custom Field", fq)
        if existing:
            LOG.info("Custom Field exists: %s", fq)
            continue
        LOG.info("Creating Custom Field: %s", fq)
        try:
            fc.create("Custom Field", cf)
        except FrappeError as e:
            LOG.error("Custom Field %s failed: %s", fq, e)
            raise


# ------------------------------- main ---------------------------------------
def _locate_xlsx() -> pathlib.Path:
    for p in ATTACH_CANDIDATES:
        if pathlib.Path(p).exists():
            return pathlib.Path(p)
    raise FileNotFoundError(
        f"Could not find A-Shop xlsx in any of: {ATTACH_CANDIDATES}"
    )


def _fetch_all_lines(fc: FrappeClient) -> List[Dict[str, Any]]:
    """Paginated fetch of all Schedule Lines we care about.

    Returns list of dicts with `name, job_id, jwm_stage, jwm_status_emoji,
    jwm_current_station, jwm_priority`.
    """
    out: List[Dict[str, Any]] = []
    limit = 200
    start = 0
    fields = [
        "name",
        "job_id",
        "jwm_stage",
        "jwm_status_emoji",
        "jwm_current_station",
        "jwm_priority",
    ]
    while True:
        # Direct GET with limit_start (get_list in _frappe.py lacks start param)
        import requests
        fc._throttle()
        url = f"{fc.base_url}/api/resource/{requests.utils.quote(DT_TARGET, safe='')}"
        r = fc._sess.get(
            url,
            params={
                "fields": json.dumps(fields),
                "limit_page_length": str(limit),
                "limit_start": str(start),
                "order_by": "creation asc",
            },
            timeout=60,
        )
        r.raise_for_status()
        batch = r.json().get("data", [])
        if not batch:
            break
        out.extend(batch)
        LOG.info("Fetched %d rows (running total %d)", len(batch), len(out))
        if len(batch) < limit:
            break
        start += limit
    return out


def main() -> int:
    xlsx_path = _locate_xlsx()
    LOG.info("Using source: %s", xlsx_path)

    enrich = build_enrichment(xlsx_path)
    LOG.info("Built enrichment map: %d job_ids", len(enrich))

    fc = FrappeClient.from_env()
    LOG.info("Connected as: %s (dry=%s)", fc.ping(), fc.dry_run)

    ensure_custom_fields(fc)

    lines = _fetch_all_lines(fc)
    LOG.info("Fetched %d existing Schedule Lines", len(lines))

    patched = 0
    skipped_unchanged = 0
    skipped_no_match = 0
    failed = 0

    start_ts = _dt.datetime.now()

    for ln in lines:
        job_id = _norm_job_id(ln.get("job_id"))
        if not job_id:
            skipped_no_match += 1
            continue
        info = enrich.get(job_id)
        if not info:
            # Try a relaxed match (strip trailing .N release suffixes)
            base = job_id.split(".")[0]
            info = enrich.get(base)
        if not info:
            # Leave untouched but ensure stage is at least "uncategorized" --- skip if already set
            if ln.get("jwm_stage"):
                skipped_unchanged += 1
            else:
                payload = {
                    "jwm_stage": "uncategorized",
                    "jwm_status_emoji": "⚪",
                    "jwm_current_station": "",
                    "jwm_priority": "Low",
                }
                try:
                    fc.update(DT_TARGET, ln["name"], payload)
                    patched += 1
                except FrappeError as e:
                    LOG.error("update %s failed: %s", ln["name"], e)
                    failed += 1
            continue

        payload = {
            "jwm_stage": info["jwm_stage"],
            "jwm_status_emoji": info["jwm_status_emoji"],
            "jwm_current_station": info["jwm_current_station"],
            "jwm_priority": info["jwm_priority"],
        }
        # idempotent compare
        same = all(
            (ln.get(k) or "") == (v or "") for k, v in payload.items()
        )
        if same:
            skipped_unchanged += 1
            continue
        try:
            fc.update(DT_TARGET, ln["name"], payload)
            patched += 1
            if patched % 250 == 0:
                LOG.info(
                    "Progress: patched=%d skipped=%d failed=%d",
                    patched,
                    skipped_unchanged,
                    failed,
                )
        except FrappeError as e:
            LOG.error("update %s failed: %s", ln["name"], e)
            failed += 1

    elapsed = (_dt.datetime.now() - start_ts).total_seconds()

    # Distributions
    stage_dist: Dict[str, int] = {}
    station_dist: Dict[str, int] = {}
    for info in enrich.values():
        stage_dist[info["jwm_stage"]] = stage_dist.get(info["jwm_stage"], 0) + 1
        st = info["jwm_current_station"] or "(none)"
        station_dist[st] = station_dist.get(st, 0) + 1

    summary = {
        "source": str(xlsx_path),
        "total_lines": len(lines),
        "patched": patched,
        "skipped_unchanged": skipped_unchanged,
        "skipped_no_match": skipped_no_match,
        "failed": failed,
        "stage_distribution": dict(sorted(stage_dist.items(), key=lambda x: -x[1])),
        "station_distribution": dict(sorted(station_dist.items(), key=lambda x: -x[1])[:15]),
        "elapsed_seconds": round(elapsed, 1),
        "client_stats": fc.stats,
    }
    LOG.info("SUMMARY: %s", json.dumps(summary, indent=2, default=str))
    print(json.dumps(summary, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
